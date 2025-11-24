from openpyxl import load_workbook
import json
wb = load_workbook('Move Changes.xlsx')
sheet = wb.active
rows = list(sheet.iter_rows(values_only=True))
headers = rows[0]
out = []
for row in rows[1:]:
    if not any(row):
        continue
    entry = {
        'move': row[0],
        'bp': row[1],
        'pp': row[2],
        'accuracy': row[3],
        'effect_chance': row[4],
        'type': row[5],
        'note_move': row[7],
        'note': row[8],
    }
    out.append(entry)

with open('move_changes_dump.json', 'w', encoding='utf-8') as fh:
    json.dump(out, fh, indent=2)

print(f"Wrote {len(out)} entries to move_changes_dump.json")
